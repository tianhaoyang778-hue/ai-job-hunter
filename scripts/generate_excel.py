"""
generate_excel.py — 招聘岗位Excel生成脚本
用法：python generate_excel.py --data /tmp/jobs_all.json --output /path/to/output.xlsx
样式参考：templates/字节跳动AI岗位_推荐_.xlsx
"""
import argparse
import json
import os
import sys
from datetime import datetime
from pathlib import Path

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
except ImportError:
    import subprocess
    subprocess.run([sys.executable, "-m", "pip", "install", "openpyxl",
                    "--break-system-packages", "-q"], check=False)
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

COLOR_MAP = {"高": "C6EFCE", "中": "FFEB9C", "低": "FFC7CE"}
HEADER_BG = "1F3864"
HEADER_FG = "FFFFFF"
META_BG   = "EEF2FF"   # 浅紫（模板第一行背景）
META_FG   = "595959"   # 灰字（模板第一行字色）

COLUMNS = [
    ("序号",       6,  6),
    ("岗位名称",  20, 40),
    ("岗位地点",   8, 15),
    ("岗位职责",  30, 60),
    ("岗位要求",  30, 60),
    ("适合程度",   8, 10),
    ("匹配说明",  25, 50),
    ("详情页链接",20, 45),
]


def thin_border():
    s = Side(style="thin", color="CCCCCC")
    return Border(left=s, right=s, top=s, bottom=s)


def deduplicate(jobs):
    seen, result = set(), []
    for job in jobs:
        key = job.get("url") or job.get("title", "")
        if key and key not in seen:
            seen.add(key)
            result.append(job)
    return result


def write_excel(data, output_path):
    jobs     = deduplicate(data.get("jobs", []))
    company  = data.get("company", "")
    keywords = data.get("keywords", "")
    date     = data.get("date", datetime.now().strftime("%Y-%m-%d"))
    user_bg  = data.get("user_background", "")
    user_bg_short = user_bg[:50] + "..." if len(user_bg) > 50 else user_bg

    wb = Workbook()
    ws = wb.active
    ws.title = f"{company}AI实习岗位" if company else "岗位列表"

    num_cols = len(COLUMNS)
    last_col = get_column_letter(num_cols)

    # 第1行：浅紫背景 + 灰字斜体
    meta = f"✅ 数据来源：{company}校园招聘官网真实原文 | 关键词：{keywords} | 日期：{date}"
    if user_bg_short:
        meta += f" | 背景：{user_bg_short}"
    ws.append([meta])
    ws.merge_cells(f"A1:{last_col}1")
    ws["A1"].fill      = PatternFill("solid", fgColor="FF"+META_BG)
    ws["A1"].font      = Font(name="Arial", size=9, italic=True, color=META_FG)
    ws["A1"].alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[1].height = 21.75

    # 第2行：标题行
    ws.append([col[0] for col in COLUMNS])
    for cell in ws[2]:
        cell.fill      = PatternFill("solid", fgColor="FF"+HEADER_BG)
        cell.font      = Font(name="Arial", bold=True, color=HEADER_FG, size=11)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border    = thin_border()
    ws.row_dimensions[2].height = 25.5

    # 数据行
    col_max = {i: 0 for i in range(num_cols)}
    for idx, job in enumerate(jobs, 1):
        level = job.get("level", "中")
        row_data = [
            idx,
            job.get("title", ""),
            job.get("location", ""),
            job.get("responsibilities", ""),
            job.get("requirements", ""),
            level,
            job.get("match_note", ""),
            job.get("url", ""),
        ]
        ws.append(row_data)
        row_num = idx + 2
        ws.row_dimensions[row_num].height = 90

        for col_i, cell in enumerate(ws[row_num], 0):
            cell.font      = Font(name="Arial", size=10)
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            cell.border    = thin_border()
            val = str(row_data[col_i])[:100] if row_data[col_i] else ""
            col_max[col_i] = max(col_max[col_i], sum(2 if ord(c)>127 else 1 for c in val)/2)

        ws.cell(row=row_num, column=6).fill      = PatternFill("solid", fgColor="FF"+COLOR_MAP.get(level, "FFFFFF"))
        ws.cell(row=row_num, column=6).alignment = Alignment(horizontal="center", vertical="center")
        ws.cell(row=row_num, column=8).alignment = Alignment(vertical="top", wrap_text=False)

    # 自适应列宽
    for i, (_, min_w, max_w) in enumerate(COLUMNS, 1):
        ws.column_dimensions[get_column_letter(i)].width = max(min_w, min(int(col_max[i-1])+3, max_w))

    # 统计行：3行，COUNTIF 公式，与模板一致
    total      = len(jobs)
    data_end   = total + 2
    s          = total + 3
    for i, (label, formula) in enumerate([
        ("🟢 高适合度", f'=COUNTIF(F3:F{data_end},"高")'),
        ("🟡 中适合度", f'=COUNTIF(F3:F{data_end},"中")'),
        ("🔴 低适合度", f'=COUNTIF(F3:F{data_end},"低")'),
        ("📋 合计",     f'=COUNTA(B3:B{data_end})'),
    ]):
        r = s + i
        ws.cell(row=r, column=1).value = label
        ws.cell(row=r, column=2).value = formula
        for col in (1, 2):
            ws.cell(row=r, column=col).font = Font(name="Arial", bold=True, size=10)
        ws.row_dimensions[r].height = 18

    ws.auto_filter.ref = f"A2:{last_col}{data_end}"
    ws.freeze_panes    = "A3"

    Path(output_path).parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_path)
    print(f"✅ Excel已生成：{output_path}（共{total}个岗位）")


def merge_batch_files(batch_dir="/tmp"):
    import glob
    all_jobs, seen = [], set()
    for f in sorted(glob.glob(f"{batch_dir}/jobs_batch_*.json")):
        try:
            for job in json.load(open(f, encoding="utf-8")):
                key = job.get("url") or job.get("title", "")
                if key not in seen:
                    seen.add(key)
                    all_jobs.append(job)
        except Exception:
            continue
    return all_jobs


def main():
    p = argparse.ArgumentParser()
    p.add_argument("--data")
    p.add_argument("--output", required=True)
    p.add_argument("--merge-batches", action="store_true")
    p.add_argument("--company",  default="")
    p.add_argument("--keywords", default="")
    args = p.parse_args()

    if args.merge_batches:
        data = {"jobs": merge_batch_files(), "company": args.company, "keywords": args.keywords,
                "date": datetime.now().strftime("%Y-%m-%d")}
    elif args.data:
        with open(args.data, encoding="utf-8") as f:
            data = json.load(f)
        if args.company:  data["company"]  = args.company
        if args.keywords: data["keywords"] = args.keywords
    else:
        print("错误：请提供 --data 或 --merge-batches"); sys.exit(1)

    write_excel(data, args.output)


if __name__ == "__main__":
    main()
